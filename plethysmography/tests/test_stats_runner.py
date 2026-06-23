"""
End-to-end test that the stats runner reproduces the old xlsx layout when fed
``old_results/breathing_analysis_results.csv``: 11 sheets with the right
names and a sensible total row count.

Because the project-extension columns (``*_no_apnea``,
``apnea_mean_ms_imputed``, ``apnea_burden_s_per_min``) don't exist in the
old CSV, we alias them from the legacy columns before running stats so the
layout check still exercises the full parameter set used by the new
pipeline.
"""

from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
import pytest


def _have_old_results() -> bool:
    return Path("old_results/breathing_analysis_results.csv").exists()


def _inject_project_extension_columns(merged: pd.DataFrame) -> pd.DataFrame:
    """Add the new column names as aliases of their legacy values so the
    stats runner sees the same data under the names it now expects."""
    merged = merged.copy()
    aliases = {
        "mean_ttot_ms_no_apnea": "mean_ttot_ms",
        "mean_te_ms_no_apnea": "mean_te_ms",
        "mean_ti_ms_no_apnea": "mean_ti_ms",
        "mean_frequency_bpm_no_apnea": "mean_frequency_bpm",
        "apnea_mean_ms_imputed": "apnea_mean_ms",
    }
    for new_col, legacy_col in aliases.items():
        if legacy_col in merged.columns:
            merged[new_col] = merged[legacy_col]
    # Synthesize burden as rate * mean (s/min); zero where mean is NaN.
    # mean is in ms, so divide by 1000 to express the resulting burden in
    # seconds per minute (matching ``apnea_burden_s_per_min``).
    if {"apnea_rate_per_min", "apnea_mean_ms"}.issubset(merged.columns):
        merged["apnea_burden_s_per_min"] = (
            merged["apnea_rate_per_min"] * merged["apnea_mean_ms"] / 1000.0
        ).fillna(0.0)
    return merged


@pytest.mark.skipif(not _have_old_results(), reason="old_results CSV not present")
def test_stats_runner_matches_old_layout(tmp_path):
    import warnings; warnings.filterwarnings("ignore")
    from plethysmography.data_loading.data_log import load_data_log
    from plethysmography.stats import (
        prepare_breathing_data, run_statistics, write_stats_xlsx,
    )

    breathing = pd.read_csv("old_results/breathing_analysis_results.csv")
    log = load_data_log()
    merged = prepare_breathing_data(breathing, log)
    merged = _inject_project_extension_columns(merged)
    rows = run_statistics(merged)
    out = tmp_path / "stats.xlsx"
    write_stats_xlsx(rows, out, write_csv=False)

    import openpyxl
    wb = openpyxl.load_workbook(out, read_only=True, data_only=True)
    assert wb.sheetnames == [
        "P22 across groups ANOVA",
        "P22 across groups posthocs",
        "P19 and 22 within groups GEE",
        "P19 and 22 within posthocs",
        "HR P19 vs LR P22 t-tests",
        "HR P19s survival t-tests",
        "P22 groups and periods GEE",
        "P22 groups and periods posthoc",
        "P19 and 22 periods GEE",
        "P19 and 22 periods posthoc",
        "All Results",
    ]
    # Reference window after Section 3.1: deleting ``cov_instant_freq`` from
    # the CV_inst family removes ~100 rows; adding ``apnea_burden_s_per_min``
    # adds ~50 — net contraction relative to the previous 1080-1180 window.
    # Allow a generous bracket around the observed count to absorb
    # gating-driven post-hoc variability.
    n_data_rows = wb["All Results"].max_row - 1
    assert 900 <= n_data_rows <= 1000, (
        f"All Results count {n_data_rows} not in expected window 900-1000"
    )
